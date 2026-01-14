"""
Matching Engine Module for the Reconciliation App.
Core logic for cross-sheet matching and population.
"""
import logging
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
import io

import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)


def col_letter_to_index(letter: str) -> int:
    """Convert Excel column letter to 0-based index."""
    return column_index_from_string(letter.upper()) - 1


def index_to_col_letter(index: int) -> str:
    """Convert 0-based index to Excel column letter."""
    return get_column_letter(index + 1)


def get_column_letters_with_names(df: pd.DataFrame) -> List[Tuple[str, str]]:
    """
    Get list of (column_letter, column_name) tuples for a DataFrame.
    
    Returns:
        List of tuples like [('A', 'order_id'), ('B', 'amount'), ...]
    """
    result = []
    for idx, col_name in enumerate(df.columns):
        letter = index_to_col_letter(idx)
        result.append((letter, str(col_name)))
    return result


@dataclass
class MatchStats:
    """Statistics for matching results."""
    total_target_rows: int
    matched_rows: int
    unmatched_rows: int
    multiple_match_rows: int
    match_percentage: float
    
    def to_dict(self) -> Dict:
        return {
            'total_target_rows': self.total_target_rows,
            'matched_rows': self.matched_rows,
            'unmatched_rows': self.unmatched_rows,
            'multiple_match_rows': self.multiple_match_rows,
            'match_percentage': round(self.match_percentage, 2)
        }


class MatchingEngine:
    """
    Core matching engine for cross-sheet reconciliation operations.
    Supports matching between two sheets and populating values based on matches.
    """
    
    def __init__(self):
        """Initialize MatchingEngine."""
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def create_workbook(
        self,
        sales_df: pd.DataFrame,
        settlement_df: pd.DataFrame
    ) -> bytes:
        """
        Create a multi-sheet Excel workbook with Sales and Settlement data.
        
        Args:
            sales_df: Sales DataFrame
            settlement_df: Settlement DataFrame
            
        Returns:
            Excel file as bytes
        """
        self.logger.info("Creating multi-sheet workbook")
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            sales_df.to_excel(writer, sheet_name='Sales', index=False)
            settlement_df.to_excel(writer, sheet_name='Settlement', index=False)
        
        return buffer.getvalue()
    
    def find_matches(
        self,
        target_df: pd.DataFrame,
        target_col: str,
        source_df: pd.DataFrame,
        source_col: str,
        match_type: str = 'exact',
        fuzzy_threshold: int = 80,
        tolerance: float = 0.01
    ) -> Tuple[Dict[int, List[int]], int]:
        """
        Find matching rows between two DataFrames using optimized algorithms.
        
        Args:
            target_df: Target DataFrame (where we want to populate)
            target_col: Column name in target to match on
            source_df: Source DataFrame (where we get values from)
            source_col: Column name in source to match on
            match_type: 'exact', 'fuzzy', 'numeric_range', 'date_range'
            fuzzy_threshold: Threshold for fuzzy matching (0-100)
            tolerance: Tolerance for numeric range matching
            
        Returns:
            Tuple of (Dict mapping target row index to list of matching source row indices, total_rows)
        """
        matches = {}
        total_rows = len(target_df)
        
        if match_type == 'exact':
            # OPTIMIZED: Hash-based exact matching - O(n+m) instead of O(n*m)
            # Build a lookup dictionary from source values
            source_lookup = {}
            source_values = source_df[source_col].astype(str).str.strip().str.lower()
            
            for source_idx, source_val in enumerate(source_values):
                if pd.isna(source_val) or source_val == 'nan' or source_val == '':
                    continue
                if source_val not in source_lookup:
                    source_lookup[source_val] = []
                source_lookup[source_val].append(source_idx)
            
            # Now lookup each target value in the hash map
            target_values = target_df[target_col].astype(str).str.strip().str.lower()
            
            for target_idx, target_val in enumerate(target_values):
                if pd.isna(target_val) or target_val == 'nan' or target_val == '':
                    continue
                if target_val in source_lookup:
                    matches[target_idx] = source_lookup[target_val].copy()
        
        elif match_type == 'numeric_range':
            # For numeric range, we still need comparisons but can optimize
            try:
                target_nums = pd.to_numeric(target_df[target_col], errors='coerce')
                source_nums = pd.to_numeric(source_df[source_col], errors='coerce')
                
                # Build source index for non-null values
                valid_source = [(idx, val) for idx, val in enumerate(source_nums) if pd.notna(val)]
                
                for target_idx, t_num in enumerate(target_nums):
                    if pd.isna(t_num):
                        continue
                    
                    matching_indices = []
                    for source_idx, s_num in valid_source:
                        if t_num != 0:
                            diff_ratio = abs(t_num - s_num) / abs(t_num)
                            if diff_ratio <= tolerance:
                                matching_indices.append(source_idx)
                        elif s_num == 0:
                            matching_indices.append(source_idx)
                    
                    if matching_indices:
                        matches[target_idx] = matching_indices
            except Exception:
                pass
        
        elif match_type == 'fuzzy':
            # Fuzzy matching requires pairwise comparison, but we can batch
            target_values = target_df[target_col].astype(str).str.strip().str.lower()
            source_values = source_df[source_col].astype(str).str.strip().str.lower()
            
            # Pre-filter valid values
            valid_source = [(idx, val) for idx, val in enumerate(source_values) 
                           if pd.notna(val) and val != 'nan' and val != '']
            
            for target_idx, target_val in enumerate(target_values):
                if pd.isna(target_val) or target_val == 'nan' or target_val == '':
                    continue
                
                matching_indices = []
                for source_idx, source_val in valid_source:
                    score = fuzz.ratio(target_val, source_val)
                    if score >= fuzzy_threshold:
                        matching_indices.append(source_idx)
                
                if matching_indices:
                    matches[target_idx] = matching_indices
        
        elif match_type == 'date_range':
            try:
                target_dates = pd.to_datetime(target_df[target_col], errors='coerce')
                source_dates = pd.to_datetime(source_df[source_col], errors='coerce')
                
                valid_source = [(idx, val) for idx, val in enumerate(source_dates) if pd.notna(val)]
                
                for target_idx, t_date in enumerate(target_dates):
                    if pd.isna(t_date):
                        continue
                    
                    matching_indices = []
                    for source_idx, s_date in valid_source:
                        diff_days = abs((t_date - s_date).days)
                        if diff_days <= int(tolerance):
                            matching_indices.append(source_idx)
                    
                    if matching_indices:
                        matches[target_idx] = matching_indices
            except Exception:
                pass
        
        return matches, total_rows
    
    def populate_values(
        self,
        target_df: pd.DataFrame,
        source_df: pd.DataFrame,
        matches: Dict[int, List[int]],
        target_col_letter: str,
        source_col_name: str,
        custom_col_name: str = ''
    ) -> Tuple[pd.DataFrame, Dict[str, int]]:
        """
        Populate values from source to target based on matches.
        
        Args:
            target_df: Target DataFrame to modify
            source_df: Source DataFrame to get values from
            matches: Dict mapping target row index to source row indices
            target_col_letter: Excel column letter for target (A, B, C, ...)
            source_col_name: Column name in source DataFrame
            custom_col_name: Optional custom name for the target column
            
        Returns:
            Tuple of (modified target DataFrame, stats dict)
        """
        result_df = target_df.copy()
        target_col_idx = col_letter_to_index(target_col_letter)
        
        # Ensure we have enough columns
        while len(result_df.columns) <= target_col_idx:
            new_col_name = f"Column_{len(result_df.columns) + 1}"
            result_df[new_col_name] = np.nan
        
        # Apply custom column name if provided
        if custom_col_name:
            # Ensure uniqueness by checking existing columns
            final_name = custom_col_name
            existing_cols = list(result_df.columns)
            # Don't count the column we're replacing
            existing_cols[target_col_idx] = None
            counter = 1
            while final_name in existing_cols:
                final_name = f"{custom_col_name}_{counter}"
                counter += 1
            
            cols = list(result_df.columns)
            cols[target_col_idx] = final_name
            result_df.columns = cols
        
        target_col_name = result_df.columns[target_col_idx]
        
        stats = {
            'populated': 0,
            'summed': 0,
            'errors': 0
        }
        
        for target_idx, source_indices in matches.items():
            if len(source_indices) == 1:
                # Single match - direct copy
                value = source_df[source_col_name].iloc[source_indices[0]]
                result_df.iloc[target_idx, target_col_idx] = value
                stats['populated'] += 1
                
            elif len(source_indices) > 1:
                # Multiple matches - try to sum
                values = [source_df[source_col_name].iloc[idx] for idx in source_indices]
                
                try:
                    # Try to convert to numeric and sum
                    numeric_values = [float(v) for v in values if pd.notna(v)]
                    if numeric_values:
                        result_df.iloc[target_idx, target_col_idx] = sum(numeric_values)
                        stats['summed'] += 1
                    else:
                        result_df.iloc[target_idx, target_col_idx] = "Data not numbers"
                        stats['errors'] += 1
                except (ValueError, TypeError):
                    result_df.iloc[target_idx, target_col_idx] = "Data not numbers"
                    stats['errors'] += 1
        
        return result_df, stats
    
    def run_reconciliation(
        self,
        sales_df: pd.DataFrame,
        settlement_df: pd.DataFrame,
        match_rules: List[Dict],
        populate_rules: List[Dict],
        credit_note_df: pd.DataFrame = None,
        progress_callback: Optional[callable] = None
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, 'MatchStats']:
        """
        Run full reconciliation with matching and population.
        
        Args:
            sales_df: Sales DataFrame (can be None)
            settlement_df: Settlement DataFrame (can be None)
            match_rules: List of match rule dicts
            populate_rules: List of populate rule dicts
            credit_note_df: Optional Credit Note DataFrame
            progress_callback: Optional callback(current_row, total_rows, phase) for progress updates
            
        Returns:
            Tuple of (modified sales_df, modified settlement_df, modified credit_note_df, stats)
        """
        # Handle optional dataframes
        result_sales = sales_df.copy() if sales_df is not None else None
        result_settlement = settlement_df.copy() if settlement_df is not None else None
        result_credit_note = credit_note_df.copy() if credit_note_df is not None else None
        
        # Count total rows from all loaded files
        total_rows = sum([
            len(sales_df) if sales_df is not None else 0,
            len(settlement_df) if settlement_df is not None else 0,
            len(credit_note_df) if credit_note_df is not None else 0
        ])
        
        # Helper to check if a variable is a valid DataFrame
        def is_valid_df(df):
            return isinstance(df, pd.DataFrame)
        
        # Helper to get DataFrame by sheet name
        def get_df(sheet_name):
            if sheet_name == 'Sales':
                return result_sales
            elif sheet_name == 'Settlement':
                return result_settlement
            elif sheet_name == 'Credit Note' and is_valid_df(result_credit_note):
                return result_credit_note
            return None
        
        # Build match lookup based on rules
        all_matches_to_sales = {}  # target_idx -> source_indices
        all_matches_to_settlement = {}
        all_matches_to_credit_note = {}
        
        # Report starting
        if progress_callback:
            progress_callback(0, total_rows, "Starting matching...")
        
        for rule_idx, rule in enumerate(match_rules):
            # Support new format (sheet1/column1) with fallback to legacy (sales_column/settlement_column)
            sheet1 = rule.get('sheet1', 'Sales')
            sheet2 = rule.get('sheet2', 'Settlement')
            col1 = rule.get('column1') or rule.get('sales_column', '')
            col2 = rule.get('column2') or rule.get('settlement_column', '')
            match_type = rule.get('match_type', 'exact')
            fuzzy_threshold = rule.get('fuzzy_threshold', 80)
            tolerance = rule.get('tolerance', 0.01)
            
            df1 = get_df(sheet1)
            df2 = get_df(sheet2)
            
            # Use isinstance to check for valid DataFrame instead of direct None comparison
            if not is_valid_df(df1) or not is_valid_df(df2) or not col1 or not col2:
                continue
            
            if progress_callback:
                progress_callback(0, total_rows, f"Matching rule {rule_idx + 1}: {sheet1}.{col1} ↔ {sheet2}.{col2}")
            
            # Find matches from sheet1 perspective
            matches, _ = self.find_matches(
                df1, col1,
                df2, col2,
                match_type, fuzzy_threshold, tolerance
            )
            
            # Store matches based on sheet1
            if sheet1 == 'Sales':
                for target_idx, source_indices in matches.items():
                    if target_idx not in all_matches_to_sales:
                        all_matches_to_sales[target_idx] = []
                    all_matches_to_sales[target_idx].extend(source_indices)
            elif sheet1 == 'Settlement':
                for target_idx, source_indices in matches.items():
                    if target_idx not in all_matches_to_settlement:
                        all_matches_to_settlement[target_idx] = []
                    all_matches_to_settlement[target_idx].extend(source_indices)
            elif sheet1 == 'Credit Note':
                for target_idx, source_indices in matches.items():
                    if target_idx not in all_matches_to_credit_note:
                        all_matches_to_credit_note[target_idx] = []
                    all_matches_to_credit_note[target_idx].extend(source_indices)
            
            if progress_callback:
                progress_callback(len(all_matches_to_sales), total_rows, f"Found {len(all_matches_to_sales)} Sales matches...")
            
            # Find matches from sheet2 perspective (reverse)
            reverse_matches, _ = self.find_matches(
                df2, col2,
                df1, col1,
                match_type, fuzzy_threshold, tolerance
            )
            
            # Store reverse matches based on sheet2
            if sheet2 == 'Sales':
                for target_idx, source_indices in reverse_matches.items():
                    if target_idx not in all_matches_to_sales:
                        all_matches_to_sales[target_idx] = []
                    all_matches_to_sales[target_idx].extend(source_indices)
            elif sheet2 == 'Settlement':
                for target_idx, source_indices in reverse_matches.items():
                    if target_idx not in all_matches_to_settlement:
                        all_matches_to_settlement[target_idx] = []
                    all_matches_to_settlement[target_idx].extend(source_indices)
            elif sheet2 == 'Credit Note':
                for target_idx, source_indices in reverse_matches.items():
                    if target_idx not in all_matches_to_credit_note:
                        all_matches_to_credit_note[target_idx] = []
                    all_matches_to_credit_note[target_idx].extend(source_indices)
        
        if progress_callback:
            progress_callback(len(all_matches_to_sales), total_rows, "Matching complete. Applying population rules...")
        
        # Apply populate rules
        for rule in populate_rules:
            target_sheet = rule.get('target_sheet', 'Sales')
            target_col_letter = rule.get('target_column_letter', 'A')
            custom_col_name = rule.get('target_column_name', '')  # Custom column name
            source_sheet = rule.get('source_sheet', 'Settlement')
            source_col = rule.get('source_column', '')
            
            # Determine target DataFrame and matches
            if target_sheet == 'Sales':
                target_df = result_sales
                matches = all_matches_to_sales
            elif target_sheet == 'Settlement':
                target_df = result_settlement
                matches = all_matches_to_settlement
            elif target_sheet == 'Credit Note' and is_valid_df(result_credit_note):
                target_df = result_credit_note
                matches = all_matches_to_credit_note
            else:
                continue
            
            # Determine source DataFrame based on source_sheet
            if source_sheet == 'Sales':
                source_df = result_sales
            elif source_sheet == 'Settlement':
                source_df = result_settlement
            elif source_sheet == 'Credit Note' and is_valid_df(result_credit_note):
                source_df = result_credit_note
            else:
                continue
            
            if source_col and source_col in source_df.columns:
                modified_df, stats = self.populate_values(
                    target_df, source_df, matches,
                    target_col_letter, source_col, custom_col_name
                )
                
                if target_sheet == 'Sales':
                    result_sales = modified_df
                elif target_sheet == 'Settlement':
                    result_settlement = modified_df
                elif target_sheet == 'Credit Note':
                    result_credit_note = modified_df
        
        # Calculate stats - count matches across all sheets
        total_matches = len(all_matches_to_sales) + len(all_matches_to_settlement) + len(all_matches_to_credit_note)
        multiple_sales = sum(1 for indices in all_matches_to_sales.values() if len(indices) > 1)
        multiple_settlement = sum(1 for indices in all_matches_to_settlement.values() if len(indices) > 1)
        multiple_credit = sum(1 for indices in all_matches_to_credit_note.values() if len(indices) > 1)
        total_multiple = multiple_sales + multiple_settlement + multiple_credit
        
        match_stats = MatchStats(
            total_target_rows=total_rows,
            matched_rows=total_matches,
            unmatched_rows=total_rows - total_matches,
            multiple_match_rows=total_multiple,
            match_percentage=(total_matches / total_rows * 100) if total_rows > 0 else 0
        )
        
        return result_sales, result_settlement, result_credit_note, match_stats
    
    def export_reconciled_workbook(
        self,
        sales_df: pd.DataFrame = None,
        settlement_df: pd.DataFrame = None,
        credit_note_df: pd.DataFrame = None
    ) -> bytes:
        """
        Export reconciled data as multi-sheet Excel workbook.
        
        Args:
            sales_df: Modified Sales DataFrame (optional)
            settlement_df: Modified Settlement DataFrame (optional)
            credit_note_df: Modified Credit Note DataFrame (optional)
            
        Returns:
            Excel file as bytes
        """
        self.logger.info("Exporting reconciled workbook")
        
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            if sales_df is not None:
                sales_df.to_excel(writer, sheet_name='Sales', index=False)
            if settlement_df is not None:
                settlement_df.to_excel(writer, sheet_name='Settlement', index=False)
            if credit_note_df is not None:
                credit_note_df.to_excel(writer, sheet_name='Credit Note', index=False)
        
        return buffer.getvalue()
    
    def test_match(
        self,
        sales_df: pd.DataFrame,
        settlement_df: pd.DataFrame,
        match_rules: List[Dict],
        sample_size: int = 100
    ) -> MatchStats:
        """
        Test matching logic on a sample of data.
        
        Args:
            sales_df: Full sales DataFrame
            settlement_df: Full settlement DataFrame
            match_rules: Match rules to test
            sample_size: Number of rows to sample
            
        Returns:
            MatchStats object
        """
        # Sample data
        sales_sample = sales_df.head(sample_size)
        settlement_sample = settlement_df.head(sample_size)
        
        all_matches = {}
        
        for rule in match_rules:
            sales_col = rule.get('sales_column')
            settlement_col = rule.get('settlement_column')
            match_type = rule.get('match_type', 'exact')
            fuzzy_threshold = rule.get('fuzzy_threshold', 80)
            tolerance = rule.get('tolerance', 0.01)
            
            matches, _ = self.find_matches(
                sales_sample, sales_col,
                settlement_sample, settlement_col,
                match_type, fuzzy_threshold, tolerance
            )
            
            for target_idx, source_indices in matches.items():
                if target_idx not in all_matches:
                    all_matches[target_idx] = []
                all_matches[target_idx].extend(source_indices)
        
        total = len(sales_sample)
        matched = len(all_matches)
        multiple = sum(1 for indices in all_matches.values() if len(indices) > 1)
        
        return MatchStats(
            total_target_rows=total,
            matched_rows=matched,
            unmatched_rows=total - matched,
            multiple_match_rows=multiple,
            match_percentage=(matched / total * 100) if total > 0 else 0
        )
